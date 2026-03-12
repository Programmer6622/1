import os
import re
from pathlib import Path
from typing import List, Optional

from .errors import ParseError
from .models import Option, Question

DEFAULT_CORRECT_MARKERS = "#*+$=!"


def _parse_markers(value: str) -> set[str]:
    markers: set[str] = set()
    for ch in value:
        if ch in {" ", "\t", "\n", "\r", ",", ";", "|"}:
            continue
        markers.add(ch)
    return markers


CORRECT_MARKERS = set(DEFAULT_CORRECT_MARKERS)
CORRECT_MARKERS.update(_parse_markers(os.environ.get("MYTEST_CORRECT_MARKERS", "")))
MARKER_CLASS = "".join(re.escape(ch) for ch in sorted(CORRECT_MARKERS))
_POETRY_ENABLED = os.environ.get("PRESERVE_POETRY_LINES", "1").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}
try:
    _PDF_FORCE_COLUMNS = int(os.environ.get("PDF_COLUMNS", "0") or 0)
except ValueError:
    _PDF_FORCE_COLUMNS = 0
try:
    _PDF_MAX_COLUMNS = int(os.environ.get("PDF_MAX_COLUMNS", "4") or 4)
except ValueError:
    _PDF_MAX_COLUMNS = 4
try:
    _PDF_COLUMN_GAP_RATIO = float(os.environ.get("PDF_COLUMN_GAP_RATIO", "0.12") or 0.12)
except ValueError:
    _PDF_COLUMN_GAP_RATIO = 0.12
try:
    _PDF_MIN_COLUMN_WORD_RATIO = float(
        os.environ.get("PDF_MIN_COLUMN_WORD_RATIO", "0.08") or 0.08
    )
except ValueError:
    _PDF_MIN_COLUMN_WORD_RATIO = 0.08
_PDF_RECOVERY_MODE = os.environ.get("PDF_RECOVERY_MODE", "strict").strip().lower()
OPTION_LABEL_CLASS = "A-V\u0410-\u042f\u0401"
_CYR_TO_LAT_LABEL = {
    "\u0410": "A",
    "\u0411": "B",
    "\u0412": "V",
    "\u0413": "G",
    "\u0414": "D",
    "\u0415": "E",
    "\u0401": "E",
    "\u0416": "J",
    "\u0417": "Z",
    "\u0418": "I",
    "\u0419": "Y",
    "\u041a": "K",
    "\u041b": "L",
    "\u041c": "M",
    "\u041d": "N",
    "\u041e": "O",
    "\u041f": "P",
    "\u0420": "R",
    "\u0421": "S",
    "\u0422": "T",
    "\u0423": "U",
    "\u0424": "F",
    "\u0425": "X",
    "\u0426": "C",
    "\u0427": "C",
    "\u0428": "S",
    "\u0429": "S",
    "\u042b": "Y",
    "\u042d": "E",
    "\u042e": "U",
    "\u042f": "Y",
}

ANSWER_KEY_RE = re.compile(
    rf"(?i)\b(javob|jav|j|answer|answers|ans|correct|key|kalit|otvet|otv|resp|жавоб|ответ|отв|калит|тўғри|тогри)\b[ \t]*[:\-][ \t]*([{OPTION_LABEL_CLASS}1-8](?:[ \t]*[,;/][ \t]*[{OPTION_LABEL_CLASS}1-8])*)"
)
LINE_OPTION_RE = re.compile(
    rf"^\(?([{OPTION_LABEL_CLASS}])\)?(?:\s*([{MARKER_CLASS}])\s*|\s*[).:\-]{{1,2}}\s*|\s+)(.+)$",
    re.I,
)
INLINE_OPTION_RE = re.compile(
    rf"(?:^|[\s\(\[])([{OPTION_LABEL_CLASS}])(?:\s*([{MARKER_CLASS}])\s*|\s*[).:\-]{{1,2}}\s*)",
    re.I,
)
NUMERIC_OPTION_RE = re.compile(r"^(\d{1,2})[).]\s*\[\s*([-+*])\s*\]\s*(.+)$")
NUMERIC_OPTION_PREFIX_RE = re.compile(r"^\d{1,3}[).]\s*\[\s*([-+*])\s*\]")
NUMERIC_PLAIN_OPTION_RE = re.compile(r"^\(?(\d{1,2})\)?[).]?\s+(.+)$")
ANSWER_SECTION_HINT_ASCII_RE = re.compile(
    r"(?i)^\s*(javoblar?|javob|jav|answers?|answer key|ans(?:wer)?s?|key|kalit|otvety?|otv|resp(?:onse)?|correct|жавоблар?|жавоб|ответы?|отв|калит|тўғри|тогри)\b"
)
ANSWER_SECTION_HINT_RE = re.compile(
    r"(?i)\b(javob|jav|answer|answers|ans|key|kalit|otvet|otv|resp|correct|жавоб|жавоблар|ответ|ответы|отв|калит|тўғри|тогри)\b"
)
ANSWER_PAIR_RE = re.compile(
    rf"(?<!\d)(\d{{1,4}})\s*(?:[).:-]\s*|\s+)([{OPTION_LABEL_CLASS}1-8](?:\s*[,;/]\s*[{OPTION_LABEL_CLASS}1-8])*)\b"
)
AIKEN_ANSWER_RE = re.compile(
    rf"(?im)^\s*(?:answer|ans|javob|jav|j|otvet|otv|correct|key|kalit|жавоб|ответ|отв|калит|тўғри|тогри)\s*[:\-]\s*([{OPTION_LABEL_CLASS}1-8](?:\s*[,;/]\s*[{OPTION_LABEL_CLASS}1-8])*)\s*$"
)


def parse_questions_from_text(text: str) -> List[Question]:
    text = _normalize_text(text)
    answer_key_map, text = _extract_answer_key_map(text)
    # Prefer Aiken path when the format is clearly Aiken-like.
    # This avoids numeric parser false-positives on long A/B/C blocks.
    if _looks_like_aiken(text):
        blocks = _split_aiken_blocks(text)
        questions: List[Question] = []
        seq = 1
        for block in blocks:
            number_hint = _peek_question_number(block)
            external_answers = _get_external_answers(answer_key_map, number_hint, seq)
            try:
                q = _parse_block(block, external_answers)
            except ParseError:
                continue
            if q:
                questions.append(q)
                seq += 1
        if questions:
            _apply_answer_keys_to_questions(questions, answer_key_map)
            return questions
    if _looks_like_mytest(text):
        questions = _parse_mytest(text)
        _apply_answer_keys_to_questions(questions, answer_key_map)
        return questions
    numbered_quiz = _parse_numbered_4choice_quiz(text)
    if len(numbered_quiz) >= 5:
        _apply_answer_keys_to_questions(numbered_quiz, answer_key_map)
        return numbered_quiz
    if _looks_like_aiken(text):
        blocks = _split_aiken_blocks(text)
    else:
        blocks = _split_question_blocks(text)
    questions: List[Question] = []
    seq = 1
    for block in blocks:
        number_hint = _peek_question_number(block)
        external_answers = _get_external_answers(answer_key_map, number_hint, seq)
        try:
            q = _parse_block(block, external_answers)
        except ParseError:
            continue
        if q:
            questions.append(q)
            seq += 1
    fallback_4choice = _parse_unnumbered_4choice_quiz(text)
    fallback_blocks = _parse_unnumbered_question_blocks(text)
    fallback_starred = _parse_unlabeled_starred_quiz(text)

    def _correct_ratio(items: List[Question]) -> float:
        if not items:
            return 0.0
        correct = sum(1 for q in items if any(opt.is_correct for opt in q.options))
        return correct / len(items)
    if not questions:
        if len(fallback_starred) >= len(fallback_4choice) and len(fallback_starred) >= len(
            fallback_blocks
        ):
            questions = fallback_starred
        else:
            questions = fallback_4choice or fallback_blocks
    else:
        # If primary parsing found too few questions while unnumbered fallback found
        # significantly more, prefer fallback (common for markerless olympiad DOCX).
        best_fallback = fallback_4choice if len(fallback_4choice) >= len(fallback_blocks) else fallback_blocks
        if len(fallback_starred) > len(best_fallback):
            best_fallback = fallback_starred
        # Prefer precision-heavy starred fallback when other fallbacks are noisy
        # (many questions but almost no detected correct markers).
        if (
            len(fallback_starred) >= 10
            and _correct_ratio(fallback_starred) >= 0.8
            and _correct_ratio(best_fallback) < 0.2
        ):
            best_fallback = fallback_starred
        if (
            (len(questions) <= 2 and len(best_fallback) >= 5)
            or (len(questions) <= 5 and len(best_fallback) >= 20)
        ):
            questions = best_fallback
    if not questions:
        raise ParseError("No questions found in text.")
    _apply_answer_keys_to_questions(questions, answer_key_map)
    return questions


def parse_questions_from_xlsx(path: Path) -> List[Question]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ParseError("openpyxl is required to read Excel files.") from exc

    wb = load_workbook(path, data_only=True)
    questions: List[Question] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_idx = _detect_header_row(rows)
        header = _row_to_strings(rows[header_idx])
        mapping = _map_headers(header)
        start_row = header_idx + 1 if mapping else header_idx
        if not mapping:
            mapping = {
                "question": 0,
                "a": 1,
                "b": 2,
                "c": 3,
                "d": 4,
                "answer": 5,
            }
        for row in rows[start_row:]:
            q = _parse_row_as_question(row, mapping)
            if q:
                questions.append(q)

    if not questions:
        raise ParseError("No questions found in Excel file.")
    return questions


def parse_questions_from_pdf(path: Path) -> List[Question]:
    try:
        import pdfplumber
    except Exception as exc:
        raise ParseError("pdfplumber is required to parse PDF files.") from exc

    lines = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            highlight_rects = _collect_highlight_rects(page)
            words = page.extract_words()
            for word in words:
                word["highlight"] = _is_word_highlighted(word, highlight_rects)
            for column_words in _split_words_by_columns(words, page.width):
                lines.extend(_group_words_into_lines(column_words))

    ordered_text = "\n".join(line["text"] for line in lines if line["text"].strip())
    parsed: Optional[List[Question]] = None
    try:
        parsed = parse_questions_from_text(ordered_text)
        parsed = _dedupe_questions_by_number(parsed)
    except ParseError:
        parsed = None

    line_based = _dedupe_questions_by_number(_parse_questions_from_lines(lines))

    if parsed and not _should_use_highlight(lines, ordered_text, parsed):
        chosen = _choose_better_pdf_parse(parsed, line_based)
        return _recover_missing_numbered_questions(chosen, ordered_text)

    questions = line_based
    if not questions and parsed:
        return _recover_missing_numbered_questions(parsed, ordered_text)
    if not questions:
        raise ParseError("No questions found in PDF.")
    return _recover_missing_numbered_questions(questions, ordered_text)


def _question_quality_score(question: Question) -> tuple[int, int, int]:
    return (
        len(question.options),
        int(any(opt.is_correct for opt in question.options)),
        len(question.text.strip()),
    )


def _parse_plain_option_text(raw: str) -> Optional[Option]:
    text = raw.strip()
    if not text:
        return None
    is_correct = False
    if text[:1] in CORRECT_MARKERS:
        is_correct = True
        text = text[1:].strip()
    if text[-1:] in CORRECT_MARKERS:
        is_correct = True
        text = text[:-1].strip()
    text = text.rstrip(";").strip()
    if not text:
        return None
    return Option(text=text, is_correct=is_correct)


def _repair_embedded_option_lines(question: Question) -> Question:
    if not question.options:
        return question
    if "\n" not in question.text:
        return question
    lines = [line.strip() for line in question.text.splitlines() if line.strip()]
    if len(lines) < 2:
        return question
    stem_idx = next((idx for idx, line in enumerate(lines) if line.endswith(":")), None)
    if stem_idx is None or stem_idx >= len(lines) - 1:
        return question
    tail = lines[stem_idx + 1 :]
    if not tail:
        return question
    marker_tail = sum(1 for line in tail if line[:1] in CORRECT_MARKERS or line[-1:] in CORRECT_MARKERS)
    avg_words = sum(len(line.split()) for line in tail) / max(len(tail), 1)
    if marker_tail == 0 and not (len(tail) >= 2 and avg_words <= 8):
        return question

    moved: List[Option] = []
    for raw in tail:
        parsed = _parse_plain_option_text(raw)
        if parsed:
            moved.append(parsed)
    if not moved:
        return question

    first_existing = question.options[0]
    if (
        moved
        and not moved[-1].is_correct
        and first_existing.is_correct
        and first_existing.text
        and len(first_existing.text.split()) <= 5
    ):
        first_existing.text = f"{moved[-1].text} {first_existing.text}".strip()
        moved = moved[:-1]

    question.text = _join_question_lines(lines[: stem_idx + 1])
    question.options = moved + question.options
    return question


def _repair_single_option_with_embedded_list(question: Question) -> Question:
    if len(question.options) > 1:
        return question
    if "\n" not in question.text:
        return question
    lines = [line.strip() for line in question.text.splitlines() if line.strip()]
    if len(lines) < 4:
        return question
    stem = lines[0]
    tail = lines[1:]
    if len(tail) > 6:
        return question
    avg_words = sum(len(line.split()) for line in tail) / max(len(tail), 1)
    if avg_words > 12:
        return question
    parsed_tail: List[Option] = []
    for raw in tail:
        parsed = _parse_plain_option_text(raw)
        if parsed:
            parsed_tail.append(parsed)
    if len(parsed_tail) < 2:
        return question

    question.text = stem
    if question.options:
        question.options = parsed_tail + question.options
    else:
        question.options = parsed_tail
    return question


def _repair_empty_question_text(question: Question) -> Question:
    if question.text.strip():
        return question
    if not question.options:
        return question
    first = question.options[0].text.strip()
    if first and first.endswith("?"):
        question.text = first
        question.options = question.options[1:]
    return question


def _dedupe_questions_by_number(questions: List[Question]) -> List[Question]:
    if not questions:
        return questions
    numbered_order: List[int] = []
    best_by_number: dict[int, Question] = {}
    unnumbered: List[Question] = []
    for q in questions:
        q = _repair_embedded_option_lines(q)
        q = _repair_single_option_with_embedded_list(q)
        q = _repair_empty_question_text(q)
        if q.number is None:
            unnumbered.append(q)
            continue
        if q.number <= 0:
            continue
        if q.number > 9999:
            continue
        if q.number not in best_by_number:
            best_by_number[q.number] = q
            numbered_order.append(q.number)
            continue
        if _question_quality_score(q) > _question_quality_score(best_by_number[q.number]):
            best_by_number[q.number] = q
    deduped = [best_by_number[n] for n in numbered_order]
    deduped.extend(unnumbered)
    return deduped


def _parse_quality_score(questions: List[Question]) -> tuple[int, int, int, int]:
    if not questions:
        return (0, 0, 0, 0)
    numbered = [q.number for q in questions if q.number is not None]
    unique_numbered = len(set(numbered))
    if numbered:
        min_n, max_n = min(numbered), max(numbered)
        span = max(max_n - min_n + 1, 1)
        density = int((unique_numbered * 1000) / span)
    else:
        density = 0
    has_correct = sum(1 for q in questions if any(opt.is_correct for opt in q.options))
    enough_options = sum(1 for q in questions if len(q.options) >= 2)
    return (density, unique_numbered, len(questions), enough_options, has_correct)


def _choose_better_pdf_parse(text_based: List[Question], line_based: List[Question]) -> List[Question]:
    text_score = _parse_quality_score(text_based)
    line_score = _parse_quality_score(line_based)
    return line_based if line_score > text_score else text_based


def _build_numbered_blocks(text: str) -> dict[int, str]:
    blocks: dict[int, str] = {}
    starts: List[tuple[int, int]] = []
    for match in re.finditer(r"(?m)^\s*(\d{1,4})\s*[).:]*\s+", text):
        starts.append((match.start(), int(match.group(1))))
    for match in re.finditer(
        r"(?:^|[\s;|:.])([1-9]\d{1,3})(?=[A-Za-zÐ-Ð¯Ð°-ÑÐŽÑžÒšÒ›Ò’Ò“Ò²Ò³])",
        text,
    ):
        number = int(match.group(1))
        if number < 20:
            continue
        starts.append((match.start(1), number))

    starts = sorted(set(starts), key=lambda item: item[0])
    for idx, (start, number) in enumerate(starts):
        end = starts[idx + 1][0] if idx + 1 < len(starts) else len(text)
        block = text[start:end].strip()
        if block and number not in blocks:
            blocks[number] = block
    return blocks


def _parse_option_from_line(line: str, highlighted: bool = False) -> Optional[Option]:
    text = line.strip()
    if not text:
        return None
    is_correct = bool(highlighted)
    match = LINE_OPTION_RE.match(text)
    numeric_match = NUMERIC_PLAIN_OPTION_RE.match(text)
    if match:
        marker = match.group(2)
        text = match.group(3).strip()
        is_correct = is_correct or bool(marker)
    elif numeric_match:
        text = numeric_match.group(2).strip()
    elif re.match(rf"^\(?[{OPTION_LABEL_CLASS}]\)?[).:]{{0,2}}$", text, re.I):
        text = text[0]
    elif re.match(r"^\(?\d{1,2}\)?[).]?$", text):
        text = re.sub(r"[().]", "", text).strip()
    else:
        return None

    if text[:1] in CORRECT_MARKERS:
        is_correct = True
        text = text[1:].strip()
    if text[-1:] in CORRECT_MARKERS:
        is_correct = True
        text = text[:-1].strip()
    text = text.rstrip(";").strip()
    if not text:
        return None
    return Option(text=text, is_correct=is_correct)


def _parse_block_lenient(block: str, number: int) -> Optional[Question]:
    content = re.sub(r"^\s*\d{1,4}\s*[).:]*\s*", "", block).strip()
    if not content:
        return None
    next_q_match = re.search(rf"\b{number + 1}\s*[).:]+\s+", content)
    if next_q_match:
        content = content[: next_q_match.start()].strip()
        if not content:
            return None

    if ":" in content and ";" in content:
        q_text, rest = content.split(":", 1)
        tokens = [token.strip() for token in re.split(r";|\n", rest) if token.strip()]
        options: List[Option] = []
        for token in tokens:
            opt = _parse_option_from_line(token)
            if not opt:
                cleaned = token.strip()
                if cleaned[-1:] in CORRECT_MARKERS:
                    options.append(Option(text=cleaned[:-1].strip(), is_correct=True))
                elif len(cleaned.split()) <= 12:
                    options.append(Option(text=cleaned, is_correct=False))
            else:
                options.append(opt)
        if len(options) >= 2:
            return Question(text=f"{q_text.strip()}:", options=options, number=number)

    if ":" in content:
        q_text, rest = content.split(":", 1)
        tail_lines = [line.strip() for line in rest.splitlines() if line.strip()]
        marker_lines = sum(1 for line in tail_lines if line[-1:] in CORRECT_MARKERS)
        if len(tail_lines) >= 4 and marker_lines >= 2:
            option_lines = tail_lines[:6] if "3 Ñ‚Ð°" in q_text else tail_lines
            options: List[Option] = []
            for line in option_lines:
                text = line.strip()
                is_correct = False
                if text[:1] in CORRECT_MARKERS:
                    is_correct = True
                    text = text[1:].strip()
                if text[-1:] in CORRECT_MARKERS:
                    is_correct = True
                    text = text[:-1].strip()
                if text:
                    options.append(Option(text=text, is_correct=is_correct))
            if len(options) >= 2:
                return Question(text=f"{q_text.strip()}:", options=options, number=number)

        # Plain list options with no labels/markers (common in some medical test PDFs).
        if 3 <= len(tail_lines) <= 8:
            avg_words = sum(len(line.split()) for line in tail_lines) / max(len(tail_lines), 1)
            if avg_words <= 14:
                options: List[Option] = []
                for line in tail_lines:
                    parsed = _parse_plain_option_text(line)
                    if parsed:
                        options.append(parsed)
                if len(options) >= 2:
                    return Question(text=f"{q_text.strip()}:", options=options, number=number)

    lines = [line.strip() for line in content.splitlines() if line.strip()]
    if len(lines) < 3:
        return None
    q_lines: List[str] = []
    options: List[Option] = []
    started = False
    for line in lines:
        opt = _parse_option_from_line(line)
        if opt:
            started = True
            options.append(opt)
        elif started and options:
            options[-1].text = f"{options[-1].text} {line}".strip()
        else:
            q_lines.append(line)
    if len(options) < 2:
        return None
    question_text = _join_question_lines(q_lines) if q_lines else lines[0]
    if not question_text:
        return None
    return Question(text=question_text, options=options, number=number)


def _recover_missing_numbered_questions(
    questions: List[Question], ordered_text: str
) -> List[Question]:
    if not questions:
        return questions
    numbered = [q.number for q in questions if q.number is not None]
    if not numbered:
        return questions
    min_n, max_n = min(numbered), max(numbered)
    existing_numbers = set(numbered)
    missing = [n for n in range(min_n, max_n + 1) if n not in existing_numbers]
    if not missing:
        return questions

    blocks = _build_numbered_blocks(_normalize_text(ordered_text))
    recovered: List[Question] = []
    for n in missing:
        block = _extract_block_by_adjacent_numbers(ordered_text, n)
        if not block:
            block = blocks.get(n)
        if not block:
            continue
        try:
            parsed = _parse_block(block)
        except ParseError:
            parsed = None
        if not parsed:
            parsed = _parse_block_lenient(block, n)
        allow = parsed and len(parsed.options) >= 2
        if allow and _PDF_RECOVERY_MODE != "aggressive":
            allow = _is_plausible_recovered_question(parsed)
        if allow:
            recovered.append(parsed)
    if not recovered:
        return questions
    return _dedupe_questions_by_number(questions + recovered)


def _extract_block_by_adjacent_numbers(text: str, number: int) -> Optional[str]:
    # Prefer precise n..(n+1) slicing for recovery to avoid splitting on option lines
    # that start with numbers (e.g. "40 yosh...").
    pattern = rf"(?ms)^\s*{number}[).:]*\s+.*?(?=^\s*{number + 1}[).:]*\s+)"
    match = re.search(pattern, text)
    if match:
        return match.group(0).strip()
    return None


def _is_plausible_recovered_question(question: Question) -> bool:
    text = question.text.strip()
    if len(text) < 12:
        return False
    if re.fullmatch(rf"\(?[{OPTION_LABEL_CLASS}]\)?[).:]{{0,2}}", text, re.I):
        return False
    if text.lower() in {"Ñ‚ÑžÒ“Ñ€Ð¸", "karab", "Ò›Ð°Ñ€Ð°Ð±", "Ð´Ð°Ð²Ñ€Ð¸"}:
        return False
    options = [opt.text.strip() for opt in question.options if opt.text.strip()]
    if len(options) < 2:
        return False
    label_only = sum(
        1
        for opt in options
        if re.fullmatch(rf"\(?[{OPTION_LABEL_CLASS}]\)?[).:]{{0,2}}", opt, re.I)
    )
    very_short = sum(1 for opt in options if len(opt) <= 2)
    if label_only >= max(2, len(options) // 2):
        return False
    if very_short >= max(3, int(len(options) * 0.7)):
        return False
    long_opts = sum(1 for opt in options if len(opt) >= 6)
    if long_opts == 0:
        return False
    return True


def _should_use_highlight(
    lines: List[dict], ordered_text: str, questions: List[Question]
) -> bool:
    if not any(line.get("highlight") for line in lines):
        return False
    if not questions:
        return True
    block_count = len(_split_question_blocks(_normalize_text(ordered_text)))
    if block_count >= max(10, len(questions) * 3):
        return True
    low_option_ratio = sum(1 for q in questions if len(q.options) < 2) / max(len(questions), 1)
    if low_option_ratio >= 0.5 and block_count > len(questions):
        return True
    return False


def _normalize_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\s+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(rf"\b([{OPTION_LABEL_CLASS}])\s*#\s*", r"\1#", text)
    return text.strip()


def _normalize_line_breaks(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return _join_question_lines(text.split("\n"))


def _join_question_lines(lines: List[str]) -> str:
    if not lines:
        return ""
    cleaned: List[str] = []
    for line in lines:
        stripped = re.sub(r"[ \t]+", " ", line).strip()
        if not stripped:
            if _POETRY_ENABLED and cleaned and cleaned[-1] != "":
                cleaned.append("")
            continue
        if (
            stripped.isdigit()
            and cleaned
            and len(cleaned[-1]) <= 3
            and re.search(r"[A-Za-zÐ-Ð¯Ð°-ÑÐŽÑžÒšÒ›Ò’Ò“Ò²Ò³]$", cleaned[-1])
        ):
            # PDFs can split tokens like "O2" into separate lines ("O" + "2").
            cleaned[-1] = f"{cleaned[-1]}{stripped}"
            continue
        cleaned.append(stripped)
    if not cleaned:
        return ""
    if _POETRY_ENABLED:
        return "\n".join(cleaned).strip()
    return " ".join(line for line in cleaned if line).strip()


def _looks_like_mytest(text: str) -> bool:
    header_count = len(re.findall(r"(?m)^#\s*\d+(?:\.\d+)?\b", text))
    marker_count = len(re.findall(r"(?m)^\s*[-+*]", text))
    return header_count >= 3 and marker_count >= 3


def _parse_mytest(text: str) -> List[Question]:
    # Recover from glued headers like "... ANSWER: B#347 ..." seen in noisy exports.
    text = re.sub(r"(?<!\n)(#\s*\d+\b)", r"\n\1", text)
    questions: List[Question] = []
    current: Optional[Question] = None
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        if line.startswith("#"):
            match = re.match(r"^#\s*(\d+)?\s*(.*)$", line)
            if not match:
                continue
            number = int(match.group(1)) if match.group(1) else None
            q_text = match.group(2).strip()
            current = Question(text=q_text, options=[], number=number)
            questions.append(current)
            continue
        if line[0] in {"+", "-", "*"} and current:
            is_correct = line[0] in {"+", "*"}
            opt_text = line[1:].strip()
            current.options.append(Option(text=opt_text, is_correct=is_correct))
            continue
        if current:
            labeled = LINE_OPTION_RE.match(line)
            if labeled:
                marker = labeled.group(2)
                opt_text = labeled.group(3).strip()
                current.options.append(Option(text=opt_text, is_correct=bool(marker)))
                continue
        if current and current.options:
            # Continuation lines in PDFs often wrap without a marker; keep them with the last option.
            current.options[-1].text = f"{current.options[-1].text} {line}".strip()
            continue
        if current and not current.options:
            joiner = "\n" if _POETRY_ENABLED else " "
            current.text = f"{current.text}{joiner}{line}".strip()
    return [q for q in questions if q.options]


def _looks_like_aiken(text: str) -> bool:
    answers = AIKEN_ANSWER_RE.findall(text)
    if len(answers) < 2:
        return False
    option_lines = sum(1 for line in text.splitlines() if LINE_OPTION_RE.match(line.strip()))
    return option_lines >= len(answers) * 3


def _split_aiken_blocks(text: str) -> List[str]:
    blocks: List[str] = []
    current: List[str] = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line and not current:
            continue
        if line:
            current.append(line)
        elif current:
            current.append("")
        if line and AIKEN_ANSWER_RE.match(line):
            block = "\n".join(current).strip()
            if block:
                blocks.append(block)
            current = []
    if current:
        block = "\n".join(current).strip()
        if block:
            blocks.append(block)
    return blocks


def _is_unnumbered_question_start(line: str) -> bool:
    text = line.strip()
    if not text:
        return False
    if LINE_OPTION_RE.match(text) or NUMERIC_PLAIN_OPTION_RE.match(text):
        return False
    if text[0] in {"+", "-", "*"}:
        return False
    if text.endswith(("?", ":")) and len(text.split()) >= 3:
        return True
    return False


def _looks_like_option_candidate(line: str) -> bool:
    text = line.strip()
    if not text:
        return False
    if _is_unnumbered_question_start(text):
        return False
    if LINE_OPTION_RE.match(text) or NUMERIC_PLAIN_OPTION_RE.match(text):
        return False
    if len(text.split()) > 16:
        return False
    return True


def _looks_like_question_candidate(line: str) -> bool:
    text = line.strip()
    if not text:
        return False
    if _is_unnumbered_question_start(text):
        return True
    if re.match(
        r"(?i)^(Ð¾Ð¿Ñ€ÐµÐ´ÐµÐ»Ð¸Ñ‚Ðµ|ÑƒÐºÐ°Ð¶Ð¸Ñ‚Ðµ|Ð²Ñ‹Ð±ÐµÑ€Ð¸Ñ‚Ðµ|Ð½Ð°Ð·Ð¾Ð²Ð¸Ñ‚Ðµ|ÐºÐ°ÐºÐ¾Ð¹|ÐºÐ°ÐºÐ°Ñ|ÐºÐ°ÐºÐ¸Ðµ|ÐºÑ‚Ð¾|Ñ‡Ñ‚Ð¾|ÐºÐ¾Ð³Ð´Ð°|Ð³Ð´Ðµ|ÑÐºÐ¾Ð»ÑŒÐºÐ¾)\b",
        text,
    ):
        return True
    if text.endswith(".") and len(text.split()) >= 5:
        return True
    return False


def _parse_unnumbered_4choice_quiz(text: str) -> List[Question]:
    lines = [line.strip() for line in _normalize_text(text).splitlines() if line.strip()]
    if len(lines) < 8:
        return []

    questions: List[Question] = []
    i = 0
    while i < len(lines):
        if not _looks_like_question_candidate(lines[i]):
            i += 1
            continue
        q_start = i
        found = False
        j = i + 1
        while j + 3 < len(lines):
            if (
                _looks_like_option_candidate(lines[j])
                and _looks_like_option_candidate(lines[j + 1])
                and _looks_like_option_candidate(lines[j + 2])
                and _looks_like_option_candidate(lines[j + 3])
            ):
                q_lines = lines[q_start:j]
                q_text = _join_question_lines(q_lines)
                if q_text:
                    opts: List[Option] = []
                    for raw in lines[j : j + 4]:
                        parsed = _parse_option_from_line(raw)
                        if not parsed:
                            parsed = _parse_plain_option_text(raw)
                        if parsed:
                            opts.append(parsed)
                    if len(opts) >= 2:
                        questions.append(Question(text=q_text, options=opts, number=None))
                        i = j + 4
                        found = True
                        break
            # If next strong question begins before options found, stop this candidate.
            if j > i + 1 and _looks_like_question_candidate(lines[j]):
                break
            j += 1
        if not found:
            i += 1
    return questions


def _parse_numbered_4choice_quiz(text: str) -> List[Question]:
    lines = [line.strip() for line in _normalize_text(text).splitlines() if line.strip()]
    if len(lines) < 8:
        return []

    questions: List[Question] = []
    i = 0

    def _make_option(raw_text: str, highlighted: bool = False) -> Option:
        opt_text = raw_text.strip()
        is_correct = highlighted
        if opt_text[:1] in CORRECT_MARKERS:
            is_correct = True
            opt_text = opt_text[1:].strip()
        if opt_text[-1:] in CORRECT_MARKERS:
            is_correct = True
            opt_text = opt_text[:-1].strip()
        return Option(text=opt_text, is_correct=is_correct)

    def _flush_pending_as_options(
        pending: List[tuple[int, str]],
        dest: List[Option],
    ) -> None:
        for _, raw in pending:
            dest.append(_make_option(raw))

    while i < len(lines):
        q_match = re.match(r"^(\d{1,4})[\).:]\s+(.+)$", lines[i])
        if not q_match:
            i += 1
            continue
        q_num = int(q_match.group(1))
        q_lines = [q_match.group(2).strip()]
        j = i + 1
        options: List[Option] = []
        pending_enum_stem: List[tuple[int, str]] = []

        while j < len(lines):
            opt_match = re.match(r"^(\d{1,2})([).:])\s+(.+)$", lines[j])
            if opt_match and int(opt_match.group(1)) <= 8:
                opt_idx = int(opt_match.group(1))
                opt_sep = opt_match.group(2)
                opt_text = opt_match.group(3).strip()

                # Distinguish "2. Next question ..." from real option rows.
                if (
                    opt_sep in {".", ":"}
                    and opt_idx > q_num
                    and _looks_like_question_candidate(opt_text)
                    and (len(options) >= 2 or opt_idx == q_num + 1)
                ):
                    if pending_enum_stem and not options:
                        _flush_pending_as_options(pending_enum_stem, options)
                        pending_enum_stem = []
                    break
                if (
                    opt_sep == ")"
                    and opt_idx > q_num
                    and _looks_like_question_candidate(opt_text)
                    and len(options) >= 4
                ):
                    if pending_enum_stem and not options:
                        _flush_pending_as_options(pending_enum_stem, options)
                        pending_enum_stem = []
                    break

                # Some stems contain internal numbered lines (1),2),3)) and then
                # answer options start again from 1). Keep the first sequence in
                # the stem only when we observe this restart pattern.
                if not options:
                    if pending_enum_stem:
                        last_idx = pending_enum_stem[-1][0]
                        if opt_idx == last_idx + 1:
                            pending_enum_stem.append((opt_idx, opt_text))
                            j += 1
                            continue
                        if opt_idx == 1 and len(pending_enum_stem) >= 2:
                            q_lines.extend(
                                f"{num}) {txt}" for num, txt in pending_enum_stem
                            )
                            pending_enum_stem = []
                            options.append(_make_option(opt_text))
                            j += 1
                            continue
                        _flush_pending_as_options(pending_enum_stem, options)
                        pending_enum_stem = []
                    elif q_lines and q_lines[-1].rstrip().endswith(":") and opt_idx == 1:
                        pending_enum_stem.append((opt_idx, opt_text))
                        j += 1
                        continue

                options.append(_make_option(opt_text))
                j += 1
                continue

            next_q_match = re.match(r"^(\d{1,4})[\).:]\s+(.+)$", lines[j])
            if next_q_match and int(next_q_match.group(1)) > q_num:
                next_num = int(next_q_match.group(1))
                next_text = next_q_match.group(2).strip()
                # Hard boundary for merged blocks. Keep low numbers flexible unless
                # they clearly look like a question header.
                if next_num > 8 or (
                    _looks_like_question_candidate(next_text)
                    and (next_num == q_num + 1 or len(options) >= 2)
                ):
                    if pending_enum_stem and not options:
                        _flush_pending_as_options(pending_enum_stem, options)
                        pending_enum_stem = []
                    break
            if options:
                options[-1].text = f"{options[-1].text} {lines[j]}".strip()
            else:
                q_lines.append(lines[j])
            j += 1

        if pending_enum_stem and not options:
            _flush_pending_as_options(pending_enum_stem, options)

        if len(options) >= 2:
            questions.append(
                Question(text=_join_question_lines(q_lines), options=options, number=q_num)
            )
            i = j
        else:
            i += 1
    return questions


def _parse_unnumbered_question_blocks(text: str) -> List[Question]:
    lines = [line.strip() for line in _normalize_text(text).splitlines() if line.strip()]
    if len(lines) < 5:
        return []

    starts = [idx for idx, line in enumerate(lines) if _is_unnumbered_question_start(line)]
    if not starts:
        return []

    questions: List[Question] = []
    for i, start in enumerate(starts):
        end = starts[i + 1] if i + 1 < len(starts) else len(lines)
        block = lines[start:end]
        if len(block) < 5:
            continue

        option_count = 4 if len(block) >= 5 else max(2, len(block) - 1)
        q_lines = block[:-option_count]
        option_lines = block[-option_count:]
        if not q_lines or len(option_lines) < 2:
            continue

        q_text = _join_question_lines(q_lines)
        if not q_text:
            continue
        options: List[Option] = []
        for raw in option_lines:
            parsed = _parse_option_from_line(raw)
            if not parsed:
                parsed = _parse_plain_option_text(raw)
            if parsed:
                options.append(parsed)
        if len(options) < 2:
            continue
        questions.append(Question(text=q_text, options=options, number=None))
    return questions


def _parse_unlabeled_starred_quiz(text: str) -> List[Question]:
    lines = [line.strip() for line in _normalize_text(text).splitlines() if line.strip()]
    if len(lines) < 8:
        return []

    questions: List[Question] = []

    def is_questionish(line: str) -> bool:
        t = line.strip()
        if not t:
            return False
        if LINE_OPTION_RE.match(t) or NUMERIC_PLAIN_OPTION_RE.match(t):
            return False
        if t[:1] in {"+", "-", "*"}:
            return False
        if t[-1:] in CORRECT_MARKERS:
            return False
        if t.endswith((":", "?")):
            return True
        return len(t.split()) >= 4

    i = 0
    while i < len(lines):
        q_line = lines[i]
        if not is_questionish(q_line):
            i += 1
            continue

        options: List[Option] = []
        j = i + 1
        while j < len(lines):
            current = lines[j].strip()
            if not current:
                break
            if len(options) >= 3 and is_questionish(current):
                break
            if re.match(r"^\d{1,4}[).:]\s+", current):
                break
            if current.startswith("#"):
                break

            # markerless option with explicit correctness marker in text
            # (e.g. "Yotoq yara toshishi *").
            is_correct = False
            value = current
            if value[:1] in CORRECT_MARKERS:
                is_correct = True
                value = value[1:].strip()
            if value[-1:] in CORRECT_MARKERS:
                is_correct = True
                value = value[:-1].strip()
            if not value:
                j += 1
                continue
            options.append(Option(text=value, is_correct=is_correct))
            j += 1

        if len(options) >= 3 and any(opt.is_correct for opt in options):
            questions.append(Question(text=q_line, options=options, number=None))
            i = j
        else:
            i += 1

    return questions

def _peek_question_number(block: str) -> Optional[int]:
    match = re.match(r"^\s*(\d{1,4})\s*[).:]*\s*", block)
    return int(match.group(1)) if match else None


def _extract_answer_key_map(text: str) -> tuple[dict[int, set[str]], str]:
    mapping: dict[int, set[str]] = {}
    kept_lines: List[str] = []
    hint_active = False
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            hint_active = False
            kept_lines.append(raw)
            continue
        has_hint = bool(
            ANSWER_SECTION_HINT_RE.match(line) or ANSWER_SECTION_HINT_ASCII_RE.match(line)
        )
        matches = list(ANSWER_PAIR_RE.finditer(line))
        if not matches:
            hint_active = has_hint
            kept_lines.append(raw)
            continue

        covered = sum(m.end() - m.start() for m in matches)
        coverage = covered / max(len(line), 1)
        active_hint = has_hint or hint_active
        numeric_tokens = any(any(ch.isdigit() for ch in m.group(2)) for m in matches)
        looks_like_key = (
            len(matches) >= 3
            or (coverage >= 0.6 and len(line) <= 80)
            or (active_hint and len(matches) >= 1)
        )
        if numeric_tokens and not (active_hint or len(matches) >= 4):
            looks_like_key = False
        if not looks_like_key:
            hint_active = False
            kept_lines.append(raw)
            continue

        for match in matches:
            number = int(match.group(1))
            letters = _parse_answer_letters(match.group(2))
            if not letters:
                continue
            mapping.setdefault(number, set()).update(letters)
        hint_active = active_hint

    cleaned = "\n".join(kept_lines)
    return mapping, cleaned


def _parse_answer_letters(token: str) -> List[str]:
    parts = re.split(r"[\s,;/]+", token.upper())
    return [
        part
        for part in parts
        if re.fullmatch(rf"[{OPTION_LABEL_CLASS}]", part)
        or re.fullmatch(r"(?:[1-9]|1[0-2])", part)
    ]


def _canonical_option_label(token: str) -> str:
    value = token.strip().upper()
    if len(value) != 1:
        return value
    return _CYR_TO_LAT_LABEL.get(value, value)


def _get_external_answers(
    answer_key_map: dict[int, set[str]], number: Optional[int], seq: int
) -> Optional[set[str]]:
    if not answer_key_map:
        return None
    if number is not None and number in answer_key_map:
        return answer_key_map[number]
    if seq in answer_key_map:
        return answer_key_map[seq]
    return None


def _apply_external_answers(
    options: List[Option], answers: Optional[set[str]], labels: Optional[List[str]] = None
) -> bool:
    if not answers:
        return False
    applied = False
    label_to_index: dict[str, int] = {}
    for idx, label in enumerate(labels or []):
        raw = label.upper()
        label_to_index.setdefault(raw, idx)
        label_to_index.setdefault(_canonical_option_label(raw), idx)
    for token in answers:
        token = token.upper()
        canonical_token = _canonical_option_label(token)
        idx: Optional[int] = None
        if token in label_to_index:
            idx = label_to_index[token]
        elif canonical_token in label_to_index:
            idx = label_to_index[canonical_token]
        elif token.isdigit():
            idx = int(token) - 1
        if idx is None:
            continue
        if 0 <= idx < len(options):
            options[idx].is_correct = True
            applied = True
    return applied


def _apply_answer_keys_to_questions(
    questions: List[Question], answer_key_map: dict[int, set[str]]
) -> None:
    if not answer_key_map:
        return
    for seq, question in enumerate(questions, start=1):
        if any(opt.is_correct for opt in question.options):
            continue
        answers = _get_external_answers(answer_key_map, question.number, seq)
        _apply_external_answers(question.options, answers)


def _collect_highlight_rects(page) -> List[dict]:
    rects: List[dict] = []
    annots = page.annots or []
    for annot in annots:
        data = annot.get("data", {})
        if "highlight" not in str(data.get("Subtype", "")).lower():
            continue
        quad_points = data.get("QuadPoints")
        if quad_points:
            rects.extend(_quad_points_to_rects(page, quad_points))
        else:
            rects.append(
                {
                    "x0": annot["x0"],
                    "x1": annot["x1"],
                    "top": annot["top"],
                    "bottom": annot["bottom"],
                }
            )
    return rects


def _quad_points_to_rects(page, quad_points) -> List[dict]:
    rects: List[dict] = []
    points = list(quad_points)
    for i in range(0, len(points), 8):
        xs = [points[i], points[i + 2], points[i + 4], points[i + 6]]
        ys = [points[i + 1], points[i + 3], points[i + 5], points[i + 7]]
        x0, x1 = min(xs), max(xs)
        y0, y1 = min(ys), max(ys)
        rects.append(
            {
                "x0": x0,
                "x1": x1,
                "top": page.height - y1,
                "bottom": page.height - y0,
            }
        )
    return rects


def _is_word_highlighted(word: dict, rects: List[dict]) -> bool:
    if not rects:
        return False
    wx0, wx1 = word["x0"], word["x1"]
    wtop, wbottom = word["top"], word["bottom"]
    for rect in rects:
        if wx0 <= rect["x1"] and wx1 >= rect["x0"] and wtop <= rect["bottom"] and wbottom >= rect["top"]:
            return True
    return False


def _group_words_into_lines(words: List[dict], line_tolerance: float = 2.0) -> List[dict]:
    if not words:
        return []
    lines: List[dict] = []
    words_sorted = sorted(words, key=lambda w: (w["top"], w["x0"]))
    current_top = None
    current_words: List[dict] = []

    def flush() -> None:
        if not current_words:
            return
        sorted_words = sorted(current_words, key=lambda w: w["x0"])
        text = " ".join(word["text"] for word in sorted_words)
        lines.append(
            {
                "text": text.strip(),
                "highlight": any(word.get("highlight") for word in current_words),
                "x0": min(word["x0"] for word in current_words),
                "x1": max(word["x1"] for word in current_words),
            }
        )

    for word in words_sorted:
        if current_top is None or abs(word["top"] - current_top) > line_tolerance:
            flush()
            current_words = [word]
            current_top = word["top"]
        else:
            current_words.append(word)
    flush()
    return lines


def _group_words_by_line(words: List[dict], line_tolerance: float = 2.0) -> List[List[dict]]:
    if not words:
        return []
    lines: List[List[dict]] = []
    words_sorted = sorted(words, key=lambda w: (w["top"], w["x0"]))
    current_top = None
    current_words: List[dict] = []

    def flush() -> None:
        nonlocal current_words
        if current_words:
            lines.append(current_words)
        current_words = []

    for word in words_sorted:
        if current_top is None or abs(word["top"] - current_top) > line_tolerance:
            flush()
            current_words = [word]
            current_top = word["top"]
        else:
            current_words.append(word)
    flush()
    return lines


def _collect_column_start_positions(words: List[dict], page_width: float) -> List[float]:
    positions: List[float] = []
    lines = _group_words_by_line(words)
    if not lines:
        return positions
    gap_threshold = max(page_width * 0.12, 40.0)
    for line in lines:
        if not line:
            continue
        line_sorted = sorted(line, key=lambda w: w["x0"])
        positions.append(line_sorted[0]["x0"])
        for prev, curr in zip(line_sorted, line_sorted[1:]):
            gap = curr["x0"] - prev["x1"]
            if gap > gap_threshold:
                positions.append(curr["x0"])
    return positions


def _median(values: List[float]) -> float:
    if not values:
        return 0.0
    vals = sorted(values)
    mid = len(vals) // 2
    if len(vals) % 2:
        return float(vals[mid])
    return float((vals[mid - 1] + vals[mid]) / 2.0)


def _kmeans_1d(values: List[float], k: int, iterations: int = 8) -> Optional[tuple[List[float], List[List[float]]]]:
    if k <= 1 or len(values) < k:
        return None
    values_sorted = sorted(values)
    centers = []
    for i in range(k):
        idx = int((i + 0.5) * len(values_sorted) / k)
        idx = min(max(idx, 0), len(values_sorted) - 1)
        centers.append(values_sorted[idx])

    for _ in range(iterations):
        clusters: List[List[float]] = [[] for _ in range(k)]
        for value in values_sorted:
            idx = min(range(k), key=lambda j: abs(value - centers[j]))
            clusters[idx].append(value)
        new_centers: List[Optional[float]] = []
        for cluster in clusters:
            if not cluster:
                new_centers.append(None)
            else:
                new_centers.append(sum(cluster) / len(cluster))
        if any(center is None for center in new_centers):
            return None
        if all(abs(new_centers[i] - centers[i]) < 0.5 for i in range(k)):
            centers = [float(c) for c in new_centers]
            break
        centers = [float(c) for c in new_centers]
    return centers, clusters


def _detect_column_boundaries_cluster(words: List[dict], page_width: float) -> List[float]:
    if len(words) < 30:
        return []
    xs = _collect_column_start_positions(words, page_width)
    if len(xs) < 6:
        return []
    if _PDF_MAX_COLUMNS <= 1 and _PDF_FORCE_COLUMNS < 2:
        return []
    total_words = len(xs)
    min_words = max(3, int(total_words * _PDF_MIN_COLUMN_WORD_RATIO))
    max_columns = max(2, min(4, _PDF_MAX_COLUMNS))
    forced_columns = _PDF_FORCE_COLUMNS if _PDF_FORCE_COLUMNS >= 2 else 0
    if forced_columns:
        max_columns = min(4, max(2, forced_columns))
        forced = True
    else:
        forced = False

    best_score: float = -1.0
    best_centers: Optional[List[float]] = None
    best_clusters: Optional[List[List[float]]] = None
    for k in range(2, max_columns + 1):
        if forced and k != max_columns:
            continue
        result = _kmeans_1d(xs, k)
        if not result:
            continue
        centers, clusters = result
        if any(len(cluster) < min_words for cluster in clusters):
            if forced:
                best_centers, best_clusters = centers, clusters
                best_score = 0.0
            continue
        centers_sorted = sorted(centers)
        gaps = [
            centers_sorted[i + 1] - centers_sorted[i]
            for i in range(len(centers_sorted) - 1)
        ]
        if not gaps:
            continue
        min_gap = min(gaps)
        gap_threshold = page_width * _PDF_COLUMN_GAP_RATIO
        if min_gap < gap_threshold:
            if forced:
                best_centers, best_clusters = centers, clusters
                best_score = 0.0
            continue
        spreads = []
        for cluster in clusters:
            if len(cluster) < 2:
                spreads.append(0.0)
            else:
                spreads.append(max(cluster) - min(cluster))
        max_spread = max(spreads) if spreads else 0.0
        score = min_gap / (max_spread + 1.0)
        if score > best_score:
            best_score = score
            best_centers = centers
            best_clusters = clusters

    if not best_centers:
        if forced and forced_columns:
            step = page_width / forced_columns
            return [step * i for i in range(1, forced_columns) if 0 < step * i < page_width]
        return []
    centers_sorted = sorted(best_centers)
    word_clusters: List[List[dict]] = [[] for _ in centers_sorted]
    for word in words:
        center = (word["x0"] + word["x1"]) / 2.0
        idx = min(range(len(centers_sorted)), key=lambda j: abs(center - centers_sorted[j]))
        word_clusters[idx].append(word)

    cluster_bounds: List[tuple[float, float]] = []
    for cluster in word_clusters:
        if not cluster:
            return []
        min_x0 = min(word["x0"] for word in cluster)
        max_x1 = max(word["x1"] for word in cluster)
        cluster_bounds.append((min_x0, max_x1))

    boundaries = []
    for i in range(len(cluster_bounds) - 1):
        boundary = (cluster_bounds[i][1] + cluster_bounds[i + 1][0]) / 2.0
        if 0 < boundary < page_width:
            boundaries.append(boundary)
    return boundaries


def _detect_column_boundaries_gap(words: List[dict], page_width: float) -> List[float]:
    if len(words) < 30:
        return []
    xs = sorted(word["x0"] for word in words)
    gaps = []
    for a, b in zip(xs, xs[1:]):
        gap = b - a
        if gap > 0:
            gaps.append((gap, a, b))
    if not gaps:
        return []
    median_gap = _median([gap for gap, _, _ in gaps])
    gap_threshold = max(page_width * 0.05, median_gap * 6.0)
    candidates = []
    for gap, a, b in gaps:
        boundary = (a + b) / 2.0
        if gap < gap_threshold:
            continue
        if not (page_width * 0.2 <= boundary <= page_width * 0.8):
            continue
        candidates.append((gap, boundary))
    if not candidates:
        return []
    # Keep only the largest few gaps to avoid over-splitting on noisy pages.
    candidates = sorted(candidates, key=lambda item: item[0], reverse=True)[:3]
    return sorted({boundary for _, boundary in candidates})


def _detect_column_boundaries(words: List[dict], page_width: float) -> List[float]:
    def boundary_gap_score(bounds: List[float]) -> float:
        if not bounds:
            return float("-inf")
        scores = []
        for boundary in bounds:
            left = []
            right = []
            for word in words:
                center = (word["x0"] + word["x1"]) / 2.0
                if center <= boundary:
                    left.append(word)
                else:
                    right.append(word)
            if not left or not right:
                return float("-inf")
            max_left = max(word["x1"] for word in left)
            min_right = min(word["x0"] for word in right)
            scores.append(min_right - max_left)
        return min(scores) if scores else float("-inf")

    cluster_bounds = _detect_column_boundaries_cluster(words, page_width)
    gap_bounds = _detect_column_boundaries_gap(words, page_width)

    if cluster_bounds and gap_bounds:
        cluster_score = boundary_gap_score(cluster_bounds)
        gap_score = boundary_gap_score(gap_bounds)
        if gap_score > cluster_score:
            return gap_bounds
        if cluster_score <= 0 and gap_score <= 0:
            return []
        return cluster_bounds

    if cluster_bounds:
        return cluster_bounds if boundary_gap_score(cluster_bounds) > 0 else []
    if gap_bounds:
        return gap_bounds if boundary_gap_score(gap_bounds) > 0 else []
    return []


def _split_words_by_columns(words: List[dict], page_width: float) -> List[List[dict]]:
    boundaries = _detect_column_boundaries(words, page_width)
    if not boundaries:
        return [words]
    columns: List[List[dict]] = [[] for _ in range(len(boundaries) + 1)]
    for word in words:
        center = (word["x0"] + word["x1"]) / 2.0
        idx = 0
        while idx < len(boundaries) and center > boundaries[idx]:
            idx += 1
        columns[idx].append(word)
    columns = [column for column in columns if column]
    total_words = sum(len(column) for column in columns)
    if len(columns) > 1:
        min_words = min(len(column) for column in columns)
        min_required = max(12, int(total_words * 0.08))
        if min_words < min_required:
            return [words]
    return columns


def _parse_questions_from_lines(lines: List[dict]) -> List[Question]:
    blocks: List[dict] = []
    current = None
    prev_number: Optional[int] = None
    for line in lines:
        text = line["text"].strip()
        if not text:
            continue
        match = re.match(r"^(\d{1,4})\s*[).:]*\s+(.*)$", text)
        if match:
            number = int(match.group(1))
            rest = match.group(2).strip()
            # Ignore tiny in-question enumerations (1., 2., 3.) after high-numbered questions.
            if (
                prev_number is not None
                and prev_number >= 100
                and number <= 10
                and (prev_number - number) >= 50
            ):
                if current is not None:
                    current["lines"].append(
                        {"text": text, "highlight": line["highlight"], "x0": line.get("x0")}
                    )
                continue
            if current and current.get("number") == number:
                current["lines"].append(
                    {"text": text, "highlight": line["highlight"], "x0": line.get("x0")}
                )
                continue
            if current:
                blocks.append(current)
            current = {"number": number, "lines": []}
            prev_number = number
            if rest:
                current["lines"].append(
                    {"text": rest, "highlight": line["highlight"], "x0": line.get("x0")}
                )
            continue
        if current is None:
            continue
        current["lines"].append(
            {"text": text, "highlight": line["highlight"], "x0": line.get("x0")}
        )
    if current:
        blocks.append(current)

    questions: List[Question] = []
    for block in blocks:
        question_text, options = _split_question_options(block["lines"])
        if not options:
            continue
        questions.append(Question(text=question_text, options=options, number=block["number"]))
    return questions


def _split_question_options(lines: List[dict]) -> tuple[str, List[Option]]:
    if not lines:
        return "", []
    numeric_option_indices = [
        idx
        for idx, line in enumerate(lines)
        if NUMERIC_PLAIN_OPTION_RE.match(line["text"].strip())
    ]
    option_label_indices = [
        idx
        for idx, line in enumerate(lines)
        if LINE_OPTION_RE.match(line["text"].strip())
    ]
    option_start_candidates = []
    # Prefer A/B/C-style option markers when they exist.
    # This prevents in-question enumerations like "1) ... 2) ... 3) ..."
    # from being misclassified as answer options.
    if len(option_label_indices) >= 2:
        option_start_candidates.append(option_label_indices[0])
    elif len(numeric_option_indices) >= 2:
        option_start_candidates.append(numeric_option_indices[0])
    no_marker_block = not option_start_candidates
    if option_start_candidates:
        option_start_idx = min(option_start_candidates)
        question_lines = [line["text"] for line in lines[:option_start_idx] if line["text"]]
        option_lines = lines[option_start_idx:]
    else:
        best_split_idx = _find_best_markerless_split(lines)
        if best_split_idx is not None:
            question_lines = [line["text"] for line in lines[:best_split_idx] if line["text"]]
            option_lines = lines[best_split_idx:]
        else:
            first_highlight_idx = next(
                (idx for idx, line in enumerate(lines) if line.get("highlight")), None
            )
            end_idx = None
            if first_highlight_idx is not None:
                for idx in range(first_highlight_idx - 1, -1, -1):
                    if lines[idx]["text"].rstrip().endswith((":", "?")):
                        end_idx = idx
                        break
                if end_idx is None:
                    end_idx = max(0, first_highlight_idx - 1)
            else:
                for idx, line in enumerate(lines):
                    if line["text"].rstrip().endswith((":", "?")):
                        end_idx = idx
            if end_idx is None:
                end_idx = 0
            question_lines = [line["text"] for line in lines[: end_idx + 1] if line["text"]]
            option_lines = lines[end_idx + 1 :]

    # In markerless blocks, only indented starts can begin options.
    if no_marker_block and option_lines:
        x_vals = [line.get("x0") for line in lines if line.get("x0") is not None]
        if x_vals:
            base_x = min(float(x) for x in x_vals)
            indent_threshold = 8.0
            while (
                option_lines
                and option_lines[0].get("x0") is not None
                and float(option_lines[0].get("x0")) < (base_x + indent_threshold)
                and len(option_lines) > 1
            ):
                question_lines.append(option_lines[0]["text"])
                option_lines = option_lines[1:]

    # Extend question continuation for 3-4 sentence stems:
    # consume leading non-marker lines ending with "." or ":" (max 4 lines).
    if option_lines:
        promoted: List[str] = []
        idx = 0
        max_promote = 4
        while idx < len(option_lines) and idx < max_promote:
            candidate = option_lines[idx]["text"].strip()
            if not candidate:
                break
            if LINE_OPTION_RE.match(candidate) or NUMERIC_PLAIN_OPTION_RE.match(candidate):
                break
            if not candidate.endswith((".", ":")):
                break
            if len(candidate.split()) < 2:
                break
            promoted.append(candidate)
            idx += 1
        # Keep at least two lines for options to avoid over-consuming.
        if promoted and (len(option_lines) - len(promoted) >= 2):
            question_lines.extend(promoted)
            option_lines = option_lines[len(promoted) :]

    question_lines = [line for line in question_lines if not re.fullmatch(r"\d{1,2}", line.strip())]
    question_text = _join_question_lines(question_lines)
    options: List[Option] = []
    labeled_mode = bool(option_start_candidates)
    if labeled_mode:
        current: Optional[Option] = None
        pre_option_lines: List[str] = []
        for line in option_lines:
            text = line["text"].strip()
            if not text:
                continue
            match = LINE_OPTION_RE.match(text)
            numeric_match = NUMERIC_PLAIN_OPTION_RE.match(text)
            if match:
                marker = match.group(2)
                opt_text = match.group(3).strip()
                is_correct = bool(marker) or bool(line.get("highlight"))
                if opt_text[:1] in CORRECT_MARKERS:
                    is_correct = True
                    opt_text = opt_text[1:].strip()
                if opt_text[-1:] in CORRECT_MARKERS:
                    is_correct = True
                    opt_text = opt_text[:-1].strip()
                if pre_option_lines:
                    numeric_pre_count = sum(
                        1 for raw in pre_option_lines if NUMERIC_PLAIN_OPTION_RE.match(raw.strip())
                    )
                    # If we have A/B/C options and pre-lines are numeric bullets (1),2),3)),
                    # treat them as part of the question stem, not answer options.
                    if numeric_pre_count >= 2:
                        question_lines.extend(pre_option_lines)
                    else:
                        pre_options: List[Option] = []
                        for raw in pre_option_lines:
                            parsed = _parse_option_from_line(raw)
                            if parsed:
                                pre_options.append(parsed)
                        if pre_options:
                            # If first detected option starts mid-sentence (OCR wraps),
                            # merge it with the last plain pre-option line.
                            if (
                                pre_options
                                and opt_text
                                and opt_text[0].islower()
                                and not pre_options[-1].is_correct
                            ):
                                opt_text = f"{pre_options[-1].text} {opt_text}".strip()
                                pre_options = pre_options[:-1]
                            options.extend(pre_options)
                    pre_option_lines = []
                options.append(Option(text=opt_text, is_correct=is_correct))
                current = options[-1]
            elif numeric_match:
                opt_text = numeric_match.group(2).strip()
                is_correct = bool(line.get("highlight"))
                if opt_text[:1] in CORRECT_MARKERS:
                    is_correct = True
                    opt_text = opt_text[1:].strip()
                if opt_text[-1:] in CORRECT_MARKERS:
                    is_correct = True
                    opt_text = opt_text[:-1].strip()
                if pre_option_lines:
                    pre_options: List[Option] = []
                    for raw in pre_option_lines:
                        parsed = _parse_option_from_line(raw)
                        if parsed:
                            pre_options.append(parsed)
                    options.extend(pre_options)
                    pre_option_lines = []
                options.append(Option(text=opt_text, is_correct=is_correct))
                current = options[-1]
            elif current:
                if re.match(r"^\d{1,4}[).:]\s+", text):
                    # Next question likely started inside the same block due extraction noise.
                    break
                current.text = f"{current.text} {text}".strip()
                if line.get("highlight"):
                    current.is_correct = True
            else:
                pre_option_lines.append(text)
    else:
        options = []
        x_candidates = [line.get("x0") for line in option_lines if line.get("x0") is not None]
        min_x = min(x_candidates) if x_candidates else None
        use_indent_mode = False
        if min_x is not None:
            indented_count = sum(1 for x in x_candidates if float(x) >= (min_x + 8.0))
            use_indent_mode = indented_count >= 1

        if not use_indent_mode:
            for line in option_lines:
                raw = line["text"].strip()
                if not raw:
                    continue
                is_correct = bool(line.get("highlight"))
                text = raw
                labeled = LINE_OPTION_RE.match(raw)
                numeric = NUMERIC_PLAIN_OPTION_RE.match(raw)
                if labeled:
                    marker = labeled.group(2)
                    text = labeled.group(3).strip()
                    is_correct = is_correct or bool(marker)
                elif numeric:
                    text = numeric.group(2).strip()
                if text[:1] in CORRECT_MARKERS:
                    is_correct = True
                    text = text[1:].strip()
                if text[-1:] in CORRECT_MARKERS:
                    is_correct = True
                    text = text[:-1].strip()
                options.append(Option(text=text, is_correct=is_correct))
        else:
            current: Optional[Option] = None
            for line in option_lines:
                raw = line["text"].strip()
                if not raw:
                    continue
                is_correct = bool(line.get("highlight"))
                text = raw
                labeled = LINE_OPTION_RE.match(raw)
                numeric = NUMERIC_PLAIN_OPTION_RE.match(raw)

                force_new = bool(labeled or numeric)
                if not force_new and min_x is not None and line.get("x0") is not None:
                    force_new = float(line.get("x0")) >= (min_x + 8.0)

                if labeled:
                    marker = labeled.group(2)
                    text = labeled.group(3).strip()
                    is_correct = is_correct or bool(marker)
                elif numeric:
                    text = numeric.group(2).strip()

                if text[:1] in CORRECT_MARKERS:
                    is_correct = True
                    text = text[1:].strip()
                if text[-1:] in CORRECT_MARKERS:
                    is_correct = True
                    text = text[:-1].strip()

                if force_new or current is None:
                    current = Option(text=text, is_correct=is_correct)
                    options.append(current)
                else:
                    if re.match(r"^\d{1,4}[).:]\s+", text):
                        break
                    current.text = f"{current.text} {text}".strip()
                    if is_correct:
                        current.is_correct = True
    return question_text, options


def _markerless_split_score(lines: List[dict], split_idx: int) -> float:
    if split_idx <= 0 or split_idx >= len(lines):
        return float("-inf")
    q_lines = [line["text"].strip() for line in lines[:split_idx] if line["text"].strip()]
    o_lines = [line for line in lines[split_idx:] if line["text"].strip()]
    if len(q_lines) < 1 or len(o_lines) < 2:
        return float("-inf")

    score = 0.0
    q_last = q_lines[-1]
    if q_last.endswith((":", "?")):
        score += 4.0
    if len(o_lines) >= 3:
        score += 2.0

    option_texts = [line["text"].strip() for line in o_lines]
    avg_words = sum(len(text.split()) for text in option_texts) / max(len(option_texts), 1)
    if avg_words <= 12:
        score += 2.0
    elif avg_words <= 16:
        score += 1.0
    else:
        score -= 1.0

    option_like = sum(1 for text in option_texts if _is_plain_option_line(text))
    if option_like >= 2:
        score += 2.0
    elif option_like == 1:
        score += 1.0

    marker_count = sum(
        1
        for text in option_texts
        if text[:1] in CORRECT_MARKERS or text[-1:] in CORRECT_MARKERS
    )
    if marker_count >= 1:
        score += 1.0

    all_x = [line.get("x0") for line in lines if line.get("x0") is not None]
    if all_x and o_lines and o_lines[0].get("x0") is not None:
        base_x = min(float(x) for x in all_x)
        first_opt_x = float(o_lines[0].get("x0"))
        if first_opt_x >= base_x + 8.0:
            score += 3.0
        else:
            score -= 3.0

    if len(q_lines) > 7:
        score -= 2.0
    return score


def _find_best_markerless_split(lines: List[dict]) -> Optional[int]:
    if len(lines) < 4:
        return None
    best_idx: Optional[int] = None
    best_score = float("-inf")
    for idx in range(1, len(lines)):
        score = _markerless_split_score(lines, idx)
        if score > best_score:
            best_score = score
            best_idx = idx
    if best_idx is None:
        return None
    if best_score < 3.0:
        return None
    return best_idx


def _split_question_blocks(text: str) -> List[str]:
    candidates: List[tuple[int, int]] = []
    for m in re.finditer(r"(?m)^\s*\d{1,4}\s*[).:]*\s+", text):
        line_start = text.rfind("\n", 0, m.start()) + 1
        line_end = text.find("\n", m.start())
        if line_end == -1:
            line_end = len(text)
        line = text[line_start:line_end].strip()
        if NUMERIC_OPTION_PREFIX_RE.match(line):
            continue
        num_match = re.match(r"^\s*(\d{1,4})", line)
        number = int(num_match.group(1)) if num_match else 0
        candidates.append((m.start(), number))

    starts: List[int] = []
    prev_number: Optional[int] = None
    for start, number in candidates:
        # Ignore small enumerations (1., 2., 3.) that appear inside high-numbered questions.
        if (
            prev_number is not None
            and prev_number >= 100
            and number <= 10
            and (prev_number - number) >= 50
        ):
            continue
        starts.append(start)
        prev_number = number
    if len(starts) >= 2:
        return _slice_blocks(text, starts)
    heuristic = rf"(?<!\d)(\d{{1,4}})(?=\s*[^\n]{{0,250}}?\b[{OPTION_LABEL_CLASS}][#\).:\-])"
    starts = [m.start() for m in re.finditer(heuristic, text, re.S)]
    if len(starts) >= 2:
        return _slice_blocks(text, starts)
    return [text]


def _slice_blocks(text: str, starts: List[int]) -> List[str]:
    starts = sorted(set(starts))
    blocks: List[str] = []
    for idx, start in enumerate(starts):
        end = starts[idx + 1] if idx + 1 < len(starts) else len(text)
        block = text[start:end].strip()
        if block:
            blocks.append(block)
    return blocks


def _parse_block(block: str, external_answers: Optional[set[str]] = None) -> Optional[Question]:
    if not block:
        return None
    block = block.strip()
    number = None
    match = re.match(r"^\s*(\d{1,4})\s*[).:]*\s*", block)
    if match:
        number = int(match.group(1))
        block = block[match.end() :].strip()

    answer_key = None
    key_match = ANSWER_KEY_RE.search(block)
    if key_match:
        tokens = set(_parse_answer_letters(key_match.group(2)))
        if external_answers:
            tokens.update(external_answers)
        external_answers = tokens or None
        if len(tokens) == 1:
            only = next(iter(tokens))
            if re.fullmatch(rf"[{OPTION_LABEL_CLASS}]", only):
                answer_key = only
        block = (block[: key_match.start()] + " " + block[key_match.end() :]).strip()

    numeric_based = _parse_block_numeric_options(block, number, external_answers)
    if numeric_based:
        return numeric_based

    line_based = _parse_block_line_options(block, answer_key, number, external_answers)
    if line_based:
        return line_based

    plain_line_based = _parse_block_plain_line_options(block, number, external_answers)
    if plain_line_based:
        return plain_line_based

    matches = list(INLINE_OPTION_RE.finditer(block))
    if len(matches) < 2:
        return _parse_block_with_signs(block, number, external_answers)

    question_text = _normalize_line_breaks(block[: matches[0].start()].strip())
    options: List[Option] = []
    labels: List[str] = []
    for idx, m in enumerate(matches):
        start = m.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(block)
        opt_text = block[start:end].strip()
        label = m.group(1).upper()
        is_correct = bool(m.group(2))
        if opt_text[:1] in CORRECT_MARKERS:
            is_correct = True
            opt_text = opt_text[1:].strip()
        options.append(Option(text=opt_text, is_correct=is_correct))
        labels.append(label)

    if answer_key:
        key_norm = _canonical_option_label(answer_key)
        for i, label in enumerate(labels):
            options[i].is_correct = (
                label == answer_key or _canonical_option_label(label) == key_norm
            )

    if not question_text:
        raise ParseError("Missing question text.")
    if len(options) < 2:
        raise ParseError("Not enough options found.")
    if not any(opt.is_correct for opt in options):
        _apply_external_answers(options, external_answers, labels)

    return Question(text=question_text, options=options, number=number)


def _parse_block_with_signs(
    block: str, number: Optional[int], external_answers: Optional[set[str]] = None
) -> Optional[Question]:
    lines = [line.strip() for line in block.splitlines() if line.strip()]
    if not lines:
        return None
    q_lines: List[str] = []
    options: List[Option] = []
    for line in lines:
        if line[0] in {"+", "-", "*"}:
            options.append(Option(text=line[1:].strip(), is_correct=line[0] in {"+", "*"}))
        else:
            if options:
                options[-1].text = f"{options[-1].text} {line}".strip()
            else:
                q_lines.append(line)
    if not options:
        return None
    question_text = _join_question_lines(q_lines)
    if not question_text:
        question_text = lines[0]
    if not any(opt.is_correct for opt in options):
        _apply_external_answers(options, external_answers)
    return Question(text=question_text, options=options, number=number)


def _parse_block_line_options(
    block: str,
    answer_key: Optional[str],
    number: Optional[int],
    external_answers: Optional[set[str]] = None,
) -> Optional[Question]:
    lines = [line.strip() for line in block.splitlines() if line.strip()]
    if not lines:
        return None
    question_lines: List[str] = []
    options: List[Option] = []
    labels: List[str] = []
    started = False

    for line in lines:
        match = LINE_OPTION_RE.match(line)
        if match:
            started = True
            label = match.group(1).upper()
            marker = match.group(2)
            text = match.group(3).strip()
            is_correct = bool(marker)
            if text[:1] in CORRECT_MARKERS:
                is_correct = True
                text = text[1:].strip()
            if text[-1:] in CORRECT_MARKERS:
                is_correct = True
                text = text[:-1].strip()
            options.append(Option(text=text, is_correct=is_correct))
            labels.append(label)
        elif started:
            t_line = line.strip()
            if t_line[-1:] in CORRECT_MARKERS:
                options[-1].is_correct = True
                t_line = t_line[:-1].strip()
            options[-1].text = f"{options[-1].text} {t_line}".strip()
        else:
            question_lines.append(line)

    if len(options) < 2:
        return None

    question_text = _join_question_lines(question_lines)
    if not question_text:
        return None

    if answer_key:
        key_norm = _canonical_option_label(answer_key)
        for i, label in enumerate(labels):
            options[i].is_correct = (
                label == answer_key or _canonical_option_label(label) == key_norm
            )

    if not any(opt.is_correct for opt in options):
        _apply_external_answers(options, external_answers, labels)

    return Question(text=question_text, options=options, number=number)


def _parse_block_numeric_options(
    block: str, number: Optional[int], external_answers: Optional[set[str]] = None
) -> Optional[Question]:
    lines = [line.strip() for line in block.splitlines() if line.strip()]
    if not lines:
        return None
    question_lines: List[str] = []
    options: List[Option] = []
    started = False

    for line in lines:
        match = NUMERIC_OPTION_RE.match(line)
        if match:
            started = True
            marker = match.group(2)
            text = match.group(3).strip()
            if text[-1:] in CORRECT_MARKERS:
                marker = "+"
                text = text[:-1].strip()
            options.append(Option(text=text, is_correct=marker in {"+", "*"}))
        elif started and options:
            t_line = line.strip()
            if t_line[-1:] in CORRECT_MARKERS:
                options[-1].is_correct = True
                t_line = t_line[:-1].strip()
            options[-1].text = f"{options[-1].text} {t_line}".strip()
        else:
            question_lines.append(line)

    if len(options) < 2:
        return None

    question_text = _join_question_lines(question_lines)
    if not question_text:
        return None
    if not any(opt.is_correct for opt in options):
        _apply_external_answers(options, external_answers)

    return Question(text=question_text, options=options, number=number)


def _is_plain_option_line(line: str) -> bool:
    text = line.strip()
    if not text:
        return False
    if re.match(r"^\d{1,4}\s*[).:]*\s+", text):
        return False
    if text.endswith(";"):
        return True
    if re.search(rf";\s*[{MARKER_CLASS}]$", text):
        return True
    if text[-1:] in CORRECT_MARKERS:
        return True
    if text.endswith(".") and (text[:1] in CORRECT_MARKERS or text[-1:] in CORRECT_MARKERS):
        return True
    return False


def _parse_block_plain_line_options(
    block: str, number: Optional[int], external_answers: Optional[set[str]] = None
) -> Optional[Question]:
    lines = [line.strip() for line in block.splitlines() if line.strip()]
    if len(lines) < 3:
        return None

    q_end_idx = next(
        (idx for idx, line in enumerate(lines) if line.rstrip().endswith(":")),
        None,
    )
    if q_end_idx is None:
        return None

    tail = lines[q_end_idx + 1 :]
    if len(tail) < 2:
        return None

    option_start = None
    for i in range(len(tail) - 1):
        window = tail[i : i + 4]
        option_like = sum(1 for ln in window if _is_plain_option_line(ln))
        if option_like >= 3 and _is_plain_option_line(tail[i]) and _is_plain_option_line(tail[i + 1]):
            option_start = i
            break
    if option_start is None:
        marker_lines = sum(1 for ln in tail if ln.strip()[-1:] in CORRECT_MARKERS)
        if len(tail) >= 4 and marker_lines >= 2:
            option_start = 0
    if option_start is None:
        return None

    option_lines = tail[option_start:]
    if len(option_lines) < 2:
        return None

    question_lines = lines[: q_end_idx + 1] + tail[:option_start]
    question_text = _join_question_lines(question_lines)
    if not question_text:
        return None

    options: List[Option] = []
    for raw in option_lines:
        text = raw.strip()
        is_correct = False
        if text[:1] in CORRECT_MARKERS:
            is_correct = True
            text = text[1:].strip()
        if text[-1:] in CORRECT_MARKERS:
            is_correct = True
            text = text[:-1].strip()
        text = text.rstrip(";").strip()
        if text:
            options.append(Option(text=text, is_correct=is_correct))

    if len(options) < 4 or len(options) > 6:
        return None
    semicolon_or_marker = sum(
        1
        for raw in option_lines
        if raw.rstrip().endswith(";")
        or raw.strip()[:1] in CORRECT_MARKERS
        or raw.strip()[-1:] in CORRECT_MARKERS
    )
    if semicolon_or_marker < max(2, len(options) // 2):
        return None
    if not any(opt.is_correct for opt in options):
        _apply_external_answers(options, external_answers)

    return Question(text=question_text, options=options, number=number)


def _detect_header_row(rows: List[tuple]) -> int:
    for idx in range(min(5, len(rows))):
        header = _row_to_strings(rows[idx])
        if _map_headers(header):
            return idx
    return 0


def _row_to_strings(row: tuple) -> List[str]:
    out = []
    for cell in row:
        if cell is None:
            out.append("")
        else:
            out.append(str(cell).strip().lower())
    return out


def _map_headers(header: List[str]) -> dict:
    question_aliases = {"question", "savol", "test", "savol matni", "savol matn"}
    answer_aliases = {
        "answer",
        "javob",
        "correct",
        "togri",
        "to'g'ri",
        "\u0436\u0430\u0432\u043e\u0431",
        "\u043a\u0430\u043b\u0438\u0442",
        "\u043e\u0442\u0432\u0435\u0442",
        "\u0442\u045e\u0493\u0440\u0438",
        "\u0442\u043e\u0433\u0440\u0438",
    }
    mapping: dict = {}
    for idx, name in enumerate(header):
        if not name:
            continue
        if name in question_aliases:
            mapping["question"] = idx
        if name in answer_aliases:
            mapping["answer"] = idx
        if name in {
            "a",
            "b",
            "c",
            "d",
            "e",
            "f",
            "g",
            "h",
            "i",
            "j",
            "k",
            "l",
            "m",
            "n",
            "o",
            "p",
            "q",
            "r",
            "s",
            "t",
            "u",
            "v",
        }:
            mapping[name] = idx
    return mapping


def _parse_row_as_question(row: tuple, mapping: dict) -> Optional[Question]:
    values = ["" if v is None else str(v).strip() for v in row]
    q_idx = mapping.get("question")
    if q_idx is None or q_idx >= len(values):
        return None
    q_text = values[q_idx]
    if not q_text:
        return None

    option_labels = [l for l in ["a", "b", "c", "d", "e", "f", "g", "h"] if l in mapping]
    options: List[Option] = []
    labels: List[str] = []
    for label in option_labels:
        idx = mapping[label]
        if idx >= len(values):
            continue
        text = values[idx].strip()
        if not text:
            continue
        is_correct = False
        if text[:1] in CORRECT_MARKERS:
            is_correct = True
            text = text[1:].strip()
        options.append(Option(text=text, is_correct=is_correct))
        labels.append(label.upper())

    if len(options) < 2:
        return None

    answer_idx = mapping.get("answer")
    answer_value = values[answer_idx].strip() if answer_idx is not None and answer_idx < len(values) else ""
    if answer_value:
        answer_value = answer_value.strip()
        upper = answer_value.upper()
        if upper in labels:
            for i, label in enumerate(labels):
                options[i].is_correct = label == upper
        elif answer_value.isdigit():
            pos = int(answer_value) - 1
            if 0 <= pos < len(options):
                for i in range(len(options)):
                    options[i].is_correct = i == pos
        else:
            for i, opt in enumerate(options):
                if opt.text.strip().lower() == answer_value.strip().lower():
                    for j in range(len(options)):
                        options[j].is_correct = j == i
                    break

    return Question(text=q_text, options=options)

